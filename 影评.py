import openpyxl
import requests
from bs4 import BeautifulSoup
import os
import time
import json
import pandas as pd
import matplotlib.pyplot as plt

# 保存状态信息和评论数据的文件名
STATE_FILE = 'wjd.txt'
TEMP_COMMENTS_FILE = 'wjd.json'
EXCEL_FILE = 'movie.xlsx'
SAVE_PATH = 'C:\\XXXX\\XXXXX\\XXXXX'     # 此处保存在自己指定路径中

def save_state(start, comments_count):
    with open(STATE_FILE, 'w') as f:
        f.write(f"{start}\n{comments_count}\n")

def load_state():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, 'r') as f:
            lines = f.readlines()
            if len(lines) >= 2:
                return int(lines[0].strip()), int(lines[1].strip())
    return 0, 0

def save_comments(comments):
    with open(TEMP_COMMENTS_FILE, 'w', encoding='utf-8') as f:
        json.dump(comments, f, ensure_ascii=False, indent=4)

def load_comments():
    if os.path.exists(TEMP_COMMENTS_FILE):
        with open(TEMP_COMMENTS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def convert_rating_to_number(rating_text):
    rating_dict = {
        '力荐': 50,
        '推荐': 40,
        '还行': 30,
        '较差': 20,
        '很差': 10
    }
    return rating_dict.get(rating_text, 0)

def get_movies(url, min_comments=1000):
    # 更新的请求头
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.85 Safari/537.36',
        'Referer': 'https://www.douban.com/',
        'Cookie': 'bid=your_bid_here; douban-fav-remind=1;'  # 需要替换为有效的Cookie
    }
    
    # 加载已保存的评论数据和状态
    comments = load_comments()
    start, comments_count = load_state()
    
    while comments_count < min_comments:
        response = requests.get(f"{url}&start={start}&limit=200", headers=headers)
        
        # 检查请求是否成功
        if response.status_code != 200:
            print(f"请求失败，状态码: {response.status_code}")
            break

        # 获取文本内容
        bs = BeautifulSoup(response.text, 'html.parser')
        
        # 获取评论项
        spanlist = bs.find_all('div', class_='comment')
        if not spanlist:  # 如果没有找到更多的评论，则停止循环
            print("没有找到更多评论，爬取结束")
            break
        
        for span in spanlist:
            try:
                var = span.find_all('span')[2]
                
                username = var.find('a').text  # 获取用户名
                rating_text = var.find_all('span', limit=3)[1].attrs['title']  # 获取评分文本
                rating = convert_rating_to_number(rating_text)  # 转换评分为数字
                comment_date = var.find_all('span', limit=3)[2].text.strip().split()[0]  # 获取评论时间并格式化为年月日
                
                useful = int(span.find('span', class_='votes').text)  # 获取有用数量
                
                context = span.select('p > span')[0].text  # 获取评论内容
                
                comments.append([username, rating, comment_date, useful, context])
                comments_count += 1
                
                # 检查是否已达到目标数量
                if comments_count >= min_comments:
                    break
            except (IndexError, AttributeError) as e:
                print(f"Error processing comment: {e}")
                continue
        
        # 打印当前进度
        print(f"已爬取评论数: {comments_count}")
        
        # 保存当前状态和评论数据
        save_state(start, comments_count)
        save_comments(comments)
        
        # 准备下一次请求
        start += 200
        
        # 避免频繁请求导致被封禁，加入较长的延时
        time.sleep(10)

    print("爬取数据完成。")

def process_and_save_data():
    comments = load_comments()

    if not comments:
        print("没有找到评论数据，请先爬取数据。")
        return

    # 使用 pandas 处理数据
    df = pd.DataFrame(comments, columns=['用户名', '评分', '评论时间', '有用', '评论内容'])
    
    # 评分转换成 int 类型
    df['评分'] = df['评分'].astype(int)
    
    # 对评论时间中的年进行分组
    df['年份'] = df['评论时间'].apply(lambda x: x.split('-')[0])
    
    # 统计每一年中评分的平均值
    avg_rating_per_year = df.groupby('年份')['评分'].mean().reset_index()

    # 打印每一年评分的平均值
    print(avg_rating_per_year)

    # 创建Excel文件并写入数据
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '无间道评论'
    
    # 写入评论数据
    sheet.append(['用户名', '评分', '评论时间', '有用', '评论内容'])  # 添加表头
    for item in comments:
        sheet.append(item)  # 不去掉年份列
    
    # 写入每年平均评分数据
    avg_sheet = wb.create_sheet(title='每年平均评分')
    avg_sheet.append(['年份', '平均评分'])  # 添加表头
    for row in avg_rating_per_year.itertuples(index=False):
        avg_sheet.append(row)

    # 创建数据可视化表
    visualize_data_sheet = wb.create_sheet(title='数据可视化')
    chart_file_path = create_visualization_chart(avg_rating_per_year)
    if chart_file_path:
        img = openpyxl.drawing.image.Image(chart_file_path)
        visualize_data_sheet.add_image(img, 'A1')

    # 保存文件路径
    file_path = os.path.join(SAVE_PATH, EXCEL_FILE)
    
    wb.save(file_path)
    print(f"文件已保存到: {file_path}")

def create_visualization_chart(avg_rating_per_year):
    try:
        x = avg_rating_per_year['年份']
        y = avg_rating_per_year['评分']

        # 解决中文乱码问题
        plt.rcParams['font.sans-serif'] = 'SimHei'
        # 解决负号无法显示问题
        plt.rcParams['axes.unicode_minus'] = False
        # 创建一个宽为8,高为6的画布
        flg = plt.figure(figsize=(8, 6), dpi=200)
        # 在画布上创建坐标系
        ax1 = plt.subplot()  # 表示将画板分成两行两列，取第一个区域，即左上角区域
        # ax1 = plt.subplot(121)
        ax1.plot(x, y)
        # 调整x轴竖着显示
        plt.xticks(rotation=90)
        # plt.title('每年反馈平均值', fontsize=20)
        plt.xlabel('年份', fontsize=16)
        plt.ylabel('评分', fontsize=16)
        # 每隔10年
        for a, b in zip(x, y):
            plt.text(a, b + 2, round(b, 2), ha='center', va='bottom', fontsize=10)

        plt.title('年反馈数统计')
        plt.tight_layout()

        # 保存图表为图片
        chart_file_path = 'chart.png'
        plt.savefig(chart_file_path)
        plt.close()
        return chart_file_path
    except Exception as e:
        print(f"Error creating chart: {e}")
        return None

def show_data():
    file_path = os.path.join(SAVE_PATH, EXCEL_FILE)
    if not os.path.exists(file_path):
        print("没有找到 Excel 文件，请先处理数据。")
        return

    df = pd.read_excel(file_path, sheet_name='每年平均评分')
    create_visualization_chart(df)
    print("数据展示完成。")

def main_menu():
    while True:
        print("\n请选择一个选项:")
        print("1. 爬取数据")
        print("2. 处理清洗数据")
        print("3. 展示数据")
        print("4. 退出")

        choice = input("输入选项编号: ")

        if choice == '1':
            get_movies('https://movie.douban.com/subject/1307914/comments?status=P', min_comments=1000)
        elif choice == '2':
            process_and_save_data()
        elif choice == '3':
            show_data()
        elif choice == '4':
            break
        else:
            print("无效的选项，请重新输入。")

if __name__ == '__main__':
    main_menu()
